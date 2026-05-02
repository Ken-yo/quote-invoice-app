from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def calculate_totals(items: List[Dict], tax_rate: float) -> Dict[str, int]:
    """
    明細行から小計・消費税・合計金額を計算する。
    tax_rate は 0.10 のように指定する。
    """
    subtotal = 0

    for item in items:
        quantity = int(item.get("quantity", 0))
        unit_price = int(item.get("unit_price", 0))
        subtotal += quantity * unit_price

    tax = int(subtotal * tax_rate)
    total = subtotal + tax

    return {
        "subtotal": subtotal,
        "tax": tax,
        "total": total,
    }


def create_document_no(doc_type: str, document_id: int) -> str:
    """
    帳票番号を作成する。
    """
    prefix = "QUO" if doc_type == "見積書" else "INV"
    return f"{prefix}-{document_id:05d}"


def create_excel_document(
    doc_type: str,
    document_no: str,
    issue_date: str,
    due_date: str,
    customer: Dict,
    subject: str,
    items: List[Dict],
    totals: Dict[str, int],
    note: str,
) -> bytes:
    """
    見積書・請求書をExcelファイルとして作成する。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = doc_type

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(size=20, bold=True)
    bold_font = Font(bold=True)

    ws["A1"] = doc_type
    ws["A1"].font = title_font

    ws["F2"] = "帳票番号"
    ws["G2"] = document_no
    ws["F3"] = "発行日"
    ws["G3"] = issue_date
    ws["F4"] = "支払期限" if doc_type == "請求書" else "有効期限"
    ws["G4"] = due_date

    ws["A4"] = "宛先"
    ws["A4"].font = bold_font
    ws["A5"] = customer.get("company_name", "")
    ws["A6"] = customer.get("address", "")
    ws["A7"] = f"{customer.get('contact_name', '')} 様"

    ws["A9"] = "件名"
    ws["B9"] = subject

    ws["A11"] = "下記の通り、御見積申し上げます。" if doc_type == "見積書" else "下記の通り、御請求申し上げます。"

    ws["A13"] = "合計金額"
    ws["B13"] = totals["total"]
    ws["B13"].number_format = '¥#,##0'
    ws["A13"].font = bold_font
    ws["B13"].font = Font(size=14, bold=True)

    start_row = 16

    headers = ["No", "品目", "数量", "単価", "金額"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for index, item in enumerate(items, start=1):
        row = start_row + index
        quantity = int(item.get("quantity", 0))
        unit_price = int(item.get("unit_price", 0))
        amount = quantity * unit_price

        values = [
            index,
            item.get("name", ""),
            quantity,
            unit_price,
            amount,
        ]

        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_index, value=value)
            cell.border = border

            if col_index in [4, 5]:
                cell.number_format = '¥#,##0'

    total_start_row = start_row + len(items) + 2

    ws.cell(row=total_start_row, column=4, value="小計").font = bold_font
    ws.cell(row=total_start_row, column=5, value=totals["subtotal"])
    ws.cell(row=total_start_row + 1, column=4, value="消費税").font = bold_font
    ws.cell(row=total_start_row + 1, column=5, value=totals["tax"])
    ws.cell(row=total_start_row + 2, column=4, value="合計").font = bold_font
    ws.cell(row=total_start_row + 2, column=5, value=totals["total"])

    for row in range(total_start_row, total_start_row + 3):
        ws.cell(row=row, column=5).number_format = '¥#,##0'

    note_row = total_start_row + 5
    ws.cell(row=note_row, column=1, value="備考")
    ws.cell(row=note_row, column=1).font = bold_font
    ws.cell(row=note_row + 1, column=1, value=note)

    column_widths = {
        "A": 8,
        "B": 35,
        "C": 12,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 20,
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()